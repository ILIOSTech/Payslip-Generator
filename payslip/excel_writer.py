from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU

from .money_utils import INDIAN_CURRENCY_FORMAT, amount_in_words

LOGO_PATH = Path(__file__).resolve().parent.parent / "static" / "logo.png"
ROW1_HEIGHT_PT = 42
LOGO_HEIGHT_PX = 40
LOGO_LEFT_MARGIN_PX = 4

THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_BOTTOM = Border(top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
COMPANY_FONT = Font(name="Calibri", size=20, bold=True, color="ED7D31")
ADDRESS_FONT = Font(name="Calibri", size=11, bold=True, color="2E5496")
TITLE_FONT = Font(name="Calibri", size=11, bold=True)
LABEL_FONT = Font(name="Calibri", size=10, bold=True)
VALUE_FONT = Font(name="Calibri", size=10)
NET_PAY_FONT = Font(name="Calibri", size=16, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

PAYSLIP_COLUMN_WIDTHS = [26, 16, 30, 16, 4, 4]


def _set_widths(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _merge_value(ws, row, col_start, col_end, value, font=None, alignment=None, border=None, fill=None):
    cell = ws.cell(row=row, column=col_start, value=value)
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).border = border
    if fill:
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = fill
    return cell


def _merge_block(ws, row_start, row_end, col_start, col_end, value, font=None, alignment=None, border=BOX):
    cell = ws.cell(row=row_start, column=col_start, value=value)
    ws.merge_cells(start_row=row_start, start_column=col_start, end_row=row_end, end_column=col_end)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).border = border
    return cell


def _add_logo(ws, row=0, col=0):
    """Place the logo in (row, col) (0-indexed), vertically centered within the row."""
    if not LOGO_PATH.exists():
        return
    image = XLImage(str(LOGO_PATH))
    scale = LOGO_HEIGHT_PX / image.height
    width_px = int(image.width * scale)
    height_px = LOGO_HEIGHT_PX

    row_height_px = ROW1_HEIGHT_PT * 4 / 3
    row_off_px = max(0, (row_height_px - height_px) / 2)

    marker = AnchorMarker(
        col=col, colOff=pixels_to_EMU(LOGO_LEFT_MARGIN_PX),
        row=row, rowOff=pixels_to_EMU(row_off_px),
    )
    anchor = OneCellAnchor(_from=marker)
    anchor.ext.width = pixels_to_EMU(width_px)
    anchor.ext.height = pixels_to_EMU(height_px)
    image.anchor = anchor
    ws.add_image(image)


def write_payslip_sheet(ws, company: dict, month_name: str, year: int, result):
    emp = result.employee
    _set_widths(ws, PAYSLIP_COLUMN_WIDTHS)

    ws.row_dimensions[1].height = ROW1_HEIGHT_PT
    _add_logo(ws, row=0, col=0)
    _merge_value(ws, 1, 2, 6, company["name"], COMPANY_FONT, LEFT)
    _merge_value(ws, 2, 1, 6, company["address"], ADDRESS_FONT, CENTER)
    _merge_value(ws, 3, 1, 6, f"Payslip for the Month of {month_name.upper()}, {year}",
                 TITLE_FONT, CENTER, border=TOP_BOTTOM)

    ws.row_dimensions[6].height = 18
    _merge_value(ws, 5, 1, 2, "Employee Pay Summary", LABEL_FONT, LEFT)
    _merge_value(ws, 5, 3, 4, "Employee Net Pay", LABEL_FONT, CENTER, fill=HEADER_FILL, border=BOX)

    detail_start_row = 6
    detail_rows = [
        ("Employee Name", emp.name),
        ("Designation", emp.designation),
        ("Employee ID", emp.employee_id),
        ("Employee Type", emp.employee_type),
        ("Date of Joining", emp.date_of_joining),
        ("Pay Days", f"{emp.pay_days:g} / {result.standard_days}"),
        ("UAN Number", emp.uan),
    ]
    for offset, (label, value) in enumerate(detail_rows):
        r = detail_start_row + offset
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.border = BOX
        value_cell = ws.cell(row=r, column=2, value=value)
        value_cell.font = VALUE_FONT
        value_cell.border = BOX
    detail_end_row = detail_start_row + len(detail_rows) - 1

    net_cell = _merge_block(ws, detail_start_row, detail_end_row, 3, 4, result.net, NET_PAY_FONT, CENTER)
    net_cell.number_format = INDIAN_CURRENCY_FORMAT

    header_row = detail_end_row + 2
    headers = ["EARNINGS", "AMOUNT", "DEDUCTIONS", "AMOUNT"]
    for offset, text in enumerate(headers):
        cell = ws.cell(row=header_row, column=1 + offset, value=text)
        cell.font = LABEL_FONT
        cell.fill = HEADER_FILL
        cell.border = BOX
        cell.alignment = CENTER

    earnings = [("Basic", result.basic_earned), ("Special Allowance", result.sp_allowance_earned)]
    deductions = [
        ("PF Employee's Contribution", result.pf_employee),
        ("PF Employer's Contribution", result.pf_employer),
        ("ESIC Employee's Contribution", result.esic_employee),
        ("ESIC Employer's Contribution", result.esic_employer),
        ("P.Tax", result.employee.ptax),
    ]
    body_rows = max(len(earnings), len(deductions))
    for i in range(body_rows):
        r = header_row + 1 + i
        for col in (1, 2, 3, 4):
            ws.cell(row=r, column=col).border = BOX
        if i < len(earnings):
            label, amount = earnings[i]
            ws.cell(row=r, column=1, value=label).font = VALUE_FONT
            c = ws.cell(row=r, column=2, value=amount)
            c.font = VALUE_FONT
            c.number_format = INDIAN_CURRENCY_FORMAT
        if i < len(deductions):
            label, amount = deductions[i]
            ws.cell(row=r, column=3, value=label).font = VALUE_FONT
            c = ws.cell(row=r, column=4, value=amount)
            c.font = VALUE_FONT
            c.number_format = INDIAN_CURRENCY_FORMAT

    totals_row = header_row + 1 + body_rows
    ws.cell(row=totals_row, column=1, value="Gross Earnings").font = LABEL_FONT
    gc = ws.cell(row=totals_row, column=2, value=result.gross)
    gc.font = LABEL_FONT
    gc.number_format = INDIAN_CURRENCY_FORMAT
    ws.cell(row=totals_row, column=3, value="Total Deductions").font = LABEL_FONT
    total_deductions = result.pf_employee + result.esic_employee + result.employee.ptax
    dc = ws.cell(row=totals_row, column=4, value=total_deductions)
    dc.font = LABEL_FONT
    dc.number_format = INDIAN_CURRENCY_FORMAT
    for c in range(1, 5):
        ws.cell(row=totals_row, column=c).border = BOX

    section_row = totals_row + 2
    _merge_value(ws, section_row, 1, 2, "NET PAY", LABEL_FONT, LEFT, fill=HEADER_FILL, border=BOX)
    _merge_value(ws, section_row, 3, 4, "CTC", LABEL_FONT, LEFT, fill=HEADER_FILL, border=BOX)

    def _summary_row(r, label_a, value_a, label_b, value_b, font=VALUE_FONT):
        if label_a is not None:
            ws.cell(row=r, column=1, value=label_a).font = font
            a = ws.cell(row=r, column=2, value=value_a)
            a.font = font
            a.number_format = INDIAN_CURRENCY_FORMAT
        if label_b is not None:
            ws.cell(row=r, column=3, value=label_b).font = font
            b = ws.cell(row=r, column=4, value=value_b)
            b.font = font
            b.number_format = INDIAN_CURRENCY_FORMAT
        for col in (1, 2, 3, 4):
            ws.cell(row=r, column=col).border = BOX

    total_deduction_employee = result.pf_employee + result.esic_employee + result.employee.ptax
    net_pay_items = [("Gross Earnings", result.gross), ("Total Deduction (Employee)", total_deduction_employee)]
    ctc_items = [
        ("Gross Earnings", result.gross),
        ("PF Employer's Contribution", result.pf_employer),
        ("ESIC Employer's Contribution", result.esic_employer),
    ]
    section_body_rows = max(len(net_pay_items), len(ctc_items))
    for i in range(section_body_rows):
        r = section_row + 1 + i
        label_a, value_a = net_pay_items[i] if i < len(net_pay_items) else (None, None)
        label_b, value_b = ctc_items[i] if i < len(ctc_items) else (None, None)
        _summary_row(r, label_a, value_a, label_b, value_b)

    final_row = section_row + section_body_rows + 2
    _summary_row(final_row, "Total Net Payable", result.net, "Monthly CTC", result.ctc, font=LABEL_FONT)

    words_row = final_row + 2
    words_text = f"Total Net Payable ₹{result.net:,.2f} ({amount_in_words(result.net)})"
    _merge_value(ws, words_row, 1, 6, words_text, Font(italic=True, size=10), CENTER)

    sig_row = words_row + 3
    _merge_value(ws, sig_row, 3, 4, "Authorised Signatory", LABEL_FONT, RIGHT, border=Border(top=THIN))


def write_register_sheet(ws, month_name: str, year: int, results: list):
    headers = [
        "Employee Name", "Designation", "Employee ID", "Employee Type", "Date of Joining",
        "Pay Days", "Standard Days", "UAN Number",
        "Basic", "Sp Allowance", "Gross",
        "PF Employee's Contribution", "PF Employer's Contribution",
        "ESIC Employee's Contribution", "ESIC Employer's Contribution", "P.Tax",
        "Net", "CTC",
    ]
    title = f"Salary Register - {month_name} {year}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=13, bold=True)
    title_cell.alignment = CENTER

    header_row = 2
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = LABEL_FONT
        cell.fill = HEADER_FILL
        cell.border = BOX
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(text) // 1.6)

    currency_cols = {9, 10, 11, 12, 13, 14, 15, 16, 17, 18}
    r = header_row
    for result in results:
        r += 1
        emp = result.employee
        row_values = [
            emp.name, emp.designation, emp.employee_id, emp.employee_type, emp.date_of_joining,
            emp.pay_days, result.standard_days, emp.uan,
            result.basic_earned, result.sp_allowance_earned, result.gross,
            result.pf_employee, result.pf_employer, result.esic_employee, result.esic_employer, emp.ptax,
            result.net, result.ctc,
        ]
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font = VALUE_FONT
            cell.border = BOX
            if col in currency_cols:
                cell.number_format = INDIAN_CURRENCY_FORMAT

    totals_row = r + 1
    ws.cell(row=totals_row, column=1, value="TOTAL").font = LABEL_FONT
    for col in currency_cols:
        letter = get_column_letter(col)
        cell = ws.cell(row=totals_row, column=col, value=f"=SUM({letter}{header_row + 1}:{letter}{r})")
        cell.font = LABEL_FONT
        cell.number_format = INDIAN_CURRENCY_FORMAT
        cell.border = TOP_BOTTOM
    for col in range(1, len(headers) + 1):
        if col not in currency_cols and col != 1:
            ws.cell(row=totals_row, column=col).border = TOP_BOTTOM


def build_employee_workbook(company: dict, fy_months_results: list):
    """fy_months_results: list of (month_name, year, PayslipResult) for one employee."""
    wb = Workbook()
    wb.remove(wb.active)
    for month_name, year, result in fy_months_results:
        sheet_name = f"{month_name[:3]}-{year}"
        ws = wb.create_sheet(title=sheet_name)
        write_payslip_sheet(ws, company, month_name, year, result)
    return wb


def build_register_workbook(fy_months_results: list):
    """fy_months_results: list of (month_name, year, list[PayslipResult]) for all employees."""
    wb = Workbook()
    wb.remove(wb.active)
    for month_name, year, results in fy_months_results:
        sheet_name = f"{month_name[:3]}-{year}"
        ws = wb.create_sheet(title=sheet_name)
        write_register_sheet(ws, month_name, year, results)
    return wb
